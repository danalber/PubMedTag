# -*- coding: utf-8 -*-
"""
Created on Mon Jul 23 08:48:15 2018

@author: Daniel Alber

dalber@partners.org
daniel_alber@brown.edu
"""
# json
import json

# excel
import openpyxl

# ui tools (folder/file dial)
import tkinter as tk
from tkinter import filedialog

# local directory access
import os

# list of uppercase letters
from string import ascii_uppercase

# sleep
from time import sleep

# garbage collect
import gc

# select file
def file_dial():
    
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename()
    return file_path

# select directory
def folder_dial():
    
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askdirectory()
    return file_path

# write dictionary 'dic' to path 'path'
def write_dict(path, dic):
    with open (path, 'w') as file:
        file.write(json.dumps(dic))
        file.close()
        
# read dictionary from path 'path'
def read_dict(path):
    loaded = {}
    with open (path, 'r') as file:
        loaded = json.loads(file.read())
        file.close()
    return loaded

# gen empty file
empty_file = lambda: open('tags.txt', 'w').close()

# clear screen w/ call to 'clear()'
clear = lambda: os.system('cls')

# single study w/ PMID and tags
class Study():
    def __init__(self, pmid):
        print("\nInitializing new study ...")
        self.pmid = pmid
        self.tags = []
        self.add_tags()
        
    
    
    # adds tags until 'none' entered
    def add_tags(self):
        cont = True
        while cont:
            clear()
            print("\nCurrent tags: {}".format(", ".join(self.tags)))
            not_string = True
            next_tag = ""
            while not_string:
                try:
                    next_tag = str(input
                                   ("\nEnter the next tag (or none if finished):")
                                   )
                    not_string = False
                except ValueError:
                    print("\nInvalid tag/none, try again")
            if next_tag.lower() == "none":
                cont = False
            else:
                if next_tag not in self.tags:
                    self.tags.append(next_tag)
        print("\nTags are:\n{}".format(", ".join(self.tags)))
        confirm = str(input("\nConfirm tags? (y/n)")).lower()
        if confirm == "override":
            print("admin override")
        elif confirm != "y":
            self.tags = []
            self.add_tags()
     
        
        
# main program - manages spreadsheet + behavior at top level        
class SpreadSheet():
    help_message = "\nCommands: add, save, exit, help"
    welcome_message = "\nThis will be a welcome message" 
    
    def __init__(self):
        # open file and init sheet
        self.init = False
        tries = 0
        while not self.init and tries <5:
            tries += 1
            try:
                print("\nSelect a spreadsheet to open")
                path = file_dial()
                self.wb = openpyxl.load_workbook(path)
                self.init = True
            except Exception as e:
                print(e)
                print("\nInvalid spreadsheet, try again ...\nSupported formats are: .xlsx,.xlsm,.xltx,.xltm")
                
        if self.init:
            print("\nEnter the number of the sheet you wish to edit")
            self.sheet_index = self.select_sheet()
            self.sheet_name = self.wb.sheetnames[self.sheet_index]
            self.sheet = self.wb[self.sheet_name]
            
            # init dict of tags to check against
            if str(input("\nLoad saved tags? (y/n)")).lower() != "y":
                self.unique_tags = {}
            else:
                print("\nSelect file with tags")
                path = file_dial()
                try:
                    self.unique_tags = read_dict(path)
                except Exception:
                    print("\nLoading failed ... initializing blank list of tags")
                    self.unique_tags = {}
            
            # init index of columns
            self.cols = self.generate_three_letter_list()
            self.col_ind = self.set_col_init()
            
            # keep looping
            self.keep_looping_global = True
       
    # finds first empty col index for init
    def set_col_init(self):
        auto = str(input("\nAutomatic first col init? y/n")).lower()
        if auto == "y" or auto == "yes":
            i = 0
            empty = False
            while not empty:
                if self.sheet['{}1'.format(self.cols[i])].value == None:
                    empty = True
                    break
                i += 1
        else:
            i = 13
        return i
        
    # adds study and tags to sheet
    def add_study(self, pmid):
        row = self.lookup_row(pmid)
        if row != None:
            stud = Study(pmid)
            self.keys(stud.tags)
            for tag in stud.tags:
                self.set_sheet_val(self.unique_tags[tag], row, "X")
        else:
            print("\nPMID not found, try another ... ")
             
    # initializes PMID and ensures correct length
    def init_pmid(self):
        valid = False
        error_message = "\nInvalid PMID, try again"
        pmid = 0
        while not valid:
            pmid_temp = str(input("\nWhat is the next study PMID? (or 'done' if finished entering)")).lower()
            if pmid_temp == "done":
                pmid = pmid_temp
                valid = True
            else:
                try:
                    pmid = int(pmid_temp)
                    # 8 digit pmid
                    if len(str(pmid)) <= 8:
                        valid = True
                    else:
                        print(error_message)
                except ValueError:
                    print(error_message)
        return pmid
    
    # top level loop        
    def what_to_do(self):
        print(SpreadSheet.welcome_message)
        print(SpreadSheet.help_message)
        while self.keep_looping_global:
            command = str(input("\nWhat would you like to do? (type 'help' for list of commands)")).lower()
            self.commands(command)
            clear()
        print("\nTotal tags: {}\nExiting ...".format(len(self.unique_tags.keys())))
        gc.collect()
        sleep(3)  
      
    # top level parser
    def commands(self, cmd):
        if cmd == "exit":
            yes_no = str(input("\nIs everything saved? Are you sure you want to exit? (y/n)")).lower()
            if yes_no == "y" or yes_no == "yes":
                self.keep_looping_global = False
        
        elif cmd == "add":
            # keep_adding = True
            while True:
                pmid = self.init_pmid()
                if pmid == "done":
                    break
                self.add_study(pmid)
            
        elif cmd == "save":
            self.save_sheet()
            
        elif cmd == "help":
            print(SpreadSheet.help_message)
    
        else:
            print("\nInvalid command, try again (print help for list of commands)")
        
    # helper func to set value of sheet w/ easy formatting
    def set_sheet_val(self, col, row, val):
        self.sheet['{}{}'.format(col, row)] = val
        
    # adds new keys to global dict of keys if not yet used
    def keys(self, tags):
        for tag in tags:
            if tag not in self.unique_tags:
                next_col = self.yield_next_col()
                self.unique_tags[tag] = next_col
                self.set_sheet_val(next_col,1,tag)
                
    # looks up row of study based on PMID
    def lookup_row(self, pmid):
        i = 0
        found = False
        while not found and i < 1000:
            i += 1
            if pmid == self.sheet['K{}'.format(i)].value:
                found = True
        if i == 10000:
            return None
        else:
            return i
        
    # returns next column letter index and increments index
    # (too lazy to write a proper generator)
    def yield_next_col(self):
        self.col_ind += 1
        return self.cols[self.col_ind]
        
    # select sheet from list of sheets
    def select_sheet(self):
        i = 0
        printout = []
        for sheet in self.wb.sheetnames:
            i += 1
            printout.append("{} - {}\n".format(i, sheet))
        ind_valid = True
        index = 0
        while ind_valid:
            print("\n".join(printout))
            try:
                index = int(input())
                ind_valid = False
            except ValueError:
                print("invalid index, try again")
        return index - 1
    
    # saves sheet (this is literally ripped from tsv)
    def save_sheet(self):
        print("\nWhere do you want to save the sheet?")
        directory = folder_dial()
        try:
            os.chdir(directory)
            self.wb.save(self.sheet_name + ".xlsx")
            print("\nSheet successfully saved as '{}' in: '{}'".format(
                    (self.sheet_name + ".xlsx"), directory))
        except Exception:
            self.wb.save(self.sheet_name + ".xlsx")
            print("\nError ... Sheet saved in default directory:{}".format(
                    os.getcwd()))
        empty_file()
        write_dict('tags.txt', self.unique_tags)
            
    # generates column labels A -> ZZZ (I should write a proper generator)
    def generate_three_letter_list(self):
        source = list(ascii_uppercase)
        to_return = []
        for i in source:
            to_return.append(i)
        for i in source:
            for j in source:
                to_return.append(i + j)
        for i in source:
            for j in source:
                for n in source:
                    to_return.append(i + j + n)
        return to_return

# top level code
if __name__ == "__main__":
    run = SpreadSheet()
    if run.init:
        clear()
        run.what_to_do()
    else:
        clear()
        print("\nInit failed, exiting in 3 ...")
        gc.collect()
        sleep(3)